#!/usr/bin/env python3
"""
Convierte Listado_RAF_Repetidoras Excel a NODES para data.js.
Asume Excel invertido: rx = col Rx, tx = col Tx (sin heurística).
"""
import re
import json
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Instala openpyxl: pip install openpyxl")
    raise

EXCEL_PATH = Path(__file__).resolve().parent / "Listado_RAF_Repetidoras.xlsx"
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "data.js"
DATA_JS_PATH = OUTPUT_PATH


def dms_to_decimal(s, is_south_or_west=False):
    """Convierte DMS (ej: 27° 13' 33") a grados decimales."""
    if s is None or (isinstance(s, (int, float)) and not isinstance(s, bool)):
        return float(s) if s is not None else None
    s = str(s).strip().replace(",", ".")
    # Formatos: 27° 13' 33", 27° 37′ 38" (U+2019), 70° 93° 33", 30° 1° 57,27"
    m = re.search(r"(-?\d+)\s*[^\d\w]+\s*(\d+)\s*[^\d\w]*\s*([\d.]+)", s)
    if not m:
        return None
    d, mi = int(m.group(1)), int(m.group(2))
    sec = float(m.group(3))
    if mi > 59:
        mi = mi % 60
    if sec >= 60:
        sec = sec % 60
    dec = d + mi / 60 + sec / 3600
    if is_south_or_west and dec > 0:
        dec = -dec
    return round(dec, 10)


def fmt(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d-%m-%Y")
    if isinstance(val, float):
        return str(val) if val == int(val) else str(val)
    return str(val).strip()


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    # Normalizar nombres (encoding)
    def idx(name, alt=None):
        for k, i in col.items():
            if name in k or (alt and alt in k):
                return i
        return -1

    i_nombre = idx("Nombre")
    i_signal = idx("Tx") - 1 if idx("Tx") > 0 else 3  # Señal col before Tx
    i_banda = idx("Banda")
    i_tx = idx("Tx")
    i_rx = idx("Rx")
    i_tono = idx("Tono")
    i_pot = idx("Potencia")
    i_gan = idx("Ganancia")
    i_region = idx("Regi")
    i_comuna = idx("Comuna")
    i_otorga = idx("Otorga")
    i_vence = idx("Vence")
    i_lat = idx("Latitud")
    i_lon = idx("Longitud")
    i_ubic = idx("Ubicaci", "Ubicación")

    # Cargar NODES existentes para range_km
    existing = {}
    if DATA_JS_PATH.exists():
        txt = DATA_JS_PATH.read_text(encoding="utf-8")
        match = re.search(r"const NODES = (\[.*?\]);", txt, re.DOTALL)
        if match:
            try:
                nodes = json.loads(match.group(1))
                for n in nodes:
                    existing[n.get("signal", "")] = n.get("range_km")
            except json.JSONDecodeError:
                pass

    nodes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        if len(row) <= max(i_signal, i_nombre, 0):
            continue
        signal = fmt(row[i_signal]) if i_signal >= 0 else ""
        nombre = fmt(row[i_nombre]) if i_nombre >= 0 else ""
        if not signal:
            continue

        lat_raw = row[i_lat] if i_lat >= 0 else None
        lon_raw = row[i_lon] if i_lon >= 0 else None
        lat = dms_to_decimal(lat_raw, is_south_or_west=True)
        lon = dms_to_decimal(lon_raw, is_south_or_west=True) if lon_raw else None
        if lat is None or lon is None:
            continue

        excel_tx = row[i_tx] if i_tx >= 0 else None
        excel_rx = row[i_rx] if i_rx >= 0 else None
        # Excel asumido invertido: col Tx = rx, col Rx = tx → swap siempre
        rx = fmt(excel_tx) if excel_tx is not None else ""
        tx = fmt(excel_rx) if excel_rx is not None else ""

        range_km = existing.get(signal)
        if range_km is None:
            range_km = 50.0  # default

        n = {
            "signal": signal,
            "nombre": nombre,
            "comuna": fmt(row[i_comuna]) if i_comuna >= 0 else "",
            "ubicacion": fmt(row[i_ubic]) if i_ubic >= 0 else "",
            "lat": lat,
            "lon": lon,
            "range_km": range_km,
            "potencia": fmt(row[i_pot]) if i_pot >= 0 else "",
            "ganancia": fmt(row[i_gan]) if i_gan >= 0 else "",
            "banda": fmt(row[i_banda]) if i_banda >= 0 else "",
            "rx": rx,
            "tx": tx,
            "tono": fmt(row[i_tono]) if i_tono >= 0 else "",
            "region": fmt(row[i_region]) if i_region >= 0 else "",
            "otorga": fmt(row[i_otorga]) if i_otorga >= 0 else "",
            "vence": fmt(row[i_vence]) if i_vence >= 0 else "",
        }
        nodes.append(n)

    wb.close()

    # Leer data.js para preservar VERSION y REGION_COLORS
    version = "1.6.2"
    region_colors = {}
    if DATA_JS_PATH.exists():
        txt = DATA_JS_PATH.read_text(encoding="utf-8")
        vm = re.search(r"const VERSION = '([^']+)'", txt)
        if vm:
            version = vm.group(1)
        rc = re.search(r"const REGION_COLORS = (\{[^}]+\})", txt)
        if rc:
            try:
                region_colors = json.loads(rc.group(1))
            except json.JSONDecodeError:
                pass

    out = f"""// Repetidoras Chile — datos centralizados
const VERSION = '{version}';

const NODES = {json.dumps(nodes, ensure_ascii=False)};

const REGION_COLORS = {json.dumps(region_colors, ensure_ascii=False)};
"""
    DATA_JS_PATH.write_text(out, encoding="utf-8")
    print(f"Escritos {len(nodes)} nodos en {DATA_JS_PATH}")


if __name__ == "__main__":
    main()
