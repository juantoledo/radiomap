#!/usr/bin/env python3
"""
Convierte Listado_RAF_Repetidoras Excel a data/curated_stations.csv.
Misma lógica que excel-to-nodes.py (Rx/Tx invertidos, DMS, etc.).
"""
import csv
import re
import json
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Instala openpyxl: pip install openpyxl")
    raise

EXCEL_PATH = Path(__file__).resolve().parent / "Listado_RAF_Repetidoras.xlsx"
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "data" / "curated_stations.csv"
DATA_JS_PATH = Path(__file__).resolve().parent.parent / "data" / "data.js"

CSV_HEADERS = [
    "signal", "nombre", "comuna", "ubicacion", "lat", "lon",
    "range_km", "potencia", "ganancia", "banda", "rx", "tx",
    "tono", "region", "otorga", "vence",
]


def dms_to_decimal(s, is_south_or_west=False):
    """Convierte DMS (ej: 27° 13' 33") a grados decimales."""
    if s is None or (isinstance(s, (int, float)) and not isinstance(s, bool)):
        return float(s) if s is not None else None
    s = str(s).strip().replace(",", ".")
    m = re.search(r"(-?\d+)\s*[^\d\w]+\s*(\d+)\s*[^\d\w]*\s*([\d.]+)", s)
    if not m:
        return None
    d, mi = int(m.group(1)), int(m.group(2))
    sec = float(m.group(3))
    if mi > 59:
        mi = mi % 60
    if sec >= 60:
        sec = sec % 60
    dec = d + mi / 60 + sec / 3600
    if is_south_or_west and dec > 0:
        dec = -dec
    return round(dec, 10)


def fmt(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d-%m-%Y")
    if isinstance(val, float):
        return str(val) if val == int(val) else str(val)
    return str(val).strip()


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    def idx(name, alt=None):
        for k, i in col.items():
            if name in k or (alt and alt in k):
                return i
        return -1

    i_nombre = idx("Nombre")
    i_signal = idx("Tx") - 1 if idx("Tx") > 0 else 3
    i_banda = idx("Banda")
    i_tx = idx("Tx")
    i_rx = idx("Rx")
    i_tono = idx("Tono")
    i_pot = idx("Potencia")
    i_gan = idx("Ganancia")
    i_region = idx("Regi")
    i_comuna = idx("Comuna")
    i_otorga = idx("Otorga")
    i_vence = idx("Vence")
    i_lat = idx("Latitud")
    i_lon = idx("Longitud")
    i_ubic = idx("Ubicaci", "Ubicación")

    existing_range = {}
    if DATA_JS_PATH.exists():
        txt = DATA_JS_PATH.read_text(encoding="utf-8")
        match = re.search(r"const NODES = (\[.*?\]);", txt, re.DOTALL)
        if match:
            try:
                nodes = json.loads(match.group(1))
                for n in nodes:
                    existing_range[n.get("signal", "")] = n.get("range_km")
            except json.JSONDecodeError:
                pass

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        if len(row) <= max(i_signal, i_nombre, 0):
            continue
        signal = fmt(row[i_signal]) if i_signal >= 0 else ""
        nombre = fmt(row[i_nombre]) if i_nombre >= 0 else ""
        if not signal:
            continue

        lat_raw = row[i_lat] if i_lat >= 0 else None
        lon_raw = row[i_lon] if i_lon >= 0 else None
        lat = dms_to_decimal(lat_raw, is_south_or_west=True)
        lon = dms_to_decimal(lon_raw, is_south_or_west=True) if lon_raw else None
        if lat is None or lon is None:
            continue

        excel_tx = row[i_tx] if i_tx >= 0 else None
        excel_rx = row[i_rx] if i_rx >= 0 else None
        rx = fmt(excel_tx) if excel_tx is not None else ""
        tx = fmt(excel_rx) if excel_rx is not None else ""

        range_km = existing_range.get(signal, 50.0)

        rows.append({
            "signal": signal,
            "nombre": nombre,
            "comuna": fmt(row[i_comuna]) if i_comuna >= 0 else "",
            "ubicacion": fmt(row[i_ubic]) if i_ubic >= 0 else "",
            "lat": lat,
            "lon": lon,
            "range_km": range_km,
            "potencia": fmt(row[i_pot]) if i_pot >= 0 else "",
            "ganancia": fmt(row[i_gan]) if i_gan >= 0 else "",
            "banda": fmt(row[i_banda]) if i_banda >= 0 else "",
            "rx": rx,
            "tx": tx,
            "tono": fmt(row[i_tono]) if i_tono >= 0 else "",
            "region": fmt(row[i_region]) if i_region >= 0 else "",
            "otorga": fmt(row[i_otorga]) if i_otorga >= 0 else "",
            "vence": fmt(row[i_vence]) if i_vence >= 0 else "",
        })

    wb.close()

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Generado {OUTPUT_PATH} ({len(rows)} estaciones)")


if __name__ == "__main__":
    main()
